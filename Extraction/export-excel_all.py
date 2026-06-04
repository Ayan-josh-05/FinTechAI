import os
import logging
import pandas as pd
from neo4j import GraphDatabase
from shared.config import NEO4J_URI, NEO4J_USER, NEO4J_PASSWORD

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Configure logging
logging.basicConfig(level=logging.ERROR, format='%(asctime)s | %(levelname)s | %(message)s')
logger = logging.getLogger('export_excel')

OUTPUT_FILE = "neo4j_database_export.xlsx"

import datetime
import re
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# Regex to strip illegal characters that crash openpyxl (control characters)
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def dict_to_string(d):
    """Convert dict values, lists, etc. to basic types for Excel to prevent formatting errors."""
    for k, v in d.items():
        if isinstance(v, (list, tuple)):
            clean_str = ", ".join(map(str, v))
            d[k] = ILLEGAL_CHARACTERS_RE.sub("", clean_str)
        elif isinstance(v, dict):
            clean_str = str(v)
            d[k] = ILLEGAL_CHARACTERS_RE.sub("", clean_str)
        elif isinstance(v, datetime.datetime):
            if v.tzinfo is not None:
                d[k] = v.replace(tzinfo=None)
        elif type(v).__name__ in ['DateTime', 'Date', 'Time', 'Duration']:
            d[k] = str(v)
        elif isinstance(v, str):
            d[k] = ILLEGAL_CHARACTERS_RE.sub("", v)
    return d

def format_excel_sheets(writer):
    """Applies pretty formatting to all sheets in the Excel writer."""
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        # 1. Format Headers (Row 1)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
        # 2. Format column widths and wrap text
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter # e.g. 'A', 'B', 'C'
            
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                lines = val.split('\n')
                longest_line = max([len(l) for l in lines]) if lines else 0
                if longest_line > max_length:
                    max_length = longest_line
                    
            adjusted_width = max_length + 2
            
            if adjusted_width > 60:
                adjusted_width = 60
                for cell in col[(1 if len(col) > 1 else 0):]: # Skip header for wrapping logic if we want, or do all
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                for cell in col[(1 if len(col) > 1 else 0):]:
                    cell.alignment = Alignment(vertical='top')
                    
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
        # 3. Handle pdf_paths links (clickable) and merge identical rows
        pdf_col_index = None
        for cell in worksheet[1]:
            if cell.value == 'pdf_paths':
                pdf_col_index = cell.column
                break
                
        if pdf_col_index:
            EXCEL_DIR = os.path.dirname(os.path.abspath(OUTPUT_FILE))
            
            # Set clickable hyperlinks for every row first
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=pdf_col_index)
                if cell.value:
                    paths = str(cell.value).split('\n')
                    if paths:
                        # The visible text is relative to PROJECT_ROOT (e.g. data/...). 
                        # We must resolve the hyperlink target relative to the EXCEL_DIR.
                        first_path = paths[0].strip()
                        abs_path = os.path.join(PROJECT_ROOT, first_path)
                        link_path = os.path.relpath(abs_path, EXCEL_DIR)
                        
                        cell.hyperlink = link_path
                        cell.font = Font(color="0563C1", underline="single")
            
            # Now visually merge identical blocks
            current_val = None
            start_row = 2
            for row_idx in range(2, worksheet.max_row + 2):
                cell = worksheet.cell(row=row_idx, column=pdf_col_index) if row_idx <= worksheet.max_row else None
                val = cell.value if cell else None
                
                if current_val is None:
                    current_val = val
                    start_row = row_idx
                elif val != current_val:
                    if (row_idx - 1) > start_row and current_val:
                        worksheet.merge_cells(start_row=start_row, start_column=pdf_col_index, end_row=row_idx-1, end_column=pdf_col_index)
                        merge_cell = worksheet.cell(row=start_row, column=pdf_col_index)
                        merge_cell.alignment = Alignment(vertical='center', wrap_text=True)
                    current_val = val
                    start_row = row_idx

PDF_CACHE = {}
def build_pdf_cache(project_root):
    data_dir = os.path.join(project_root, 'Extraction', 'data')
    for root, _, files in os.walk(data_dir):
        for f in files:
            if f.endswith('.pdf'):
                PDF_CACHE[f] = os.path.join(root, f)

def resolve_physical_path(storage_id):
    if not storage_id: return None
    parts = str(storage_id).split('/')
    if len(parts) >= 2:
        basename = f"{parts[-2]}_{parts[-1]}"
    else:
        basename = os.path.basename(storage_id)
    return PDF_CACHE.get(basename)

def export_all_to_excel():
    logger.info("Connecting to Neo4j database...")
    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASSWORD))
    build_pdf_cache(PROJECT_ROOT)
    
    try:
        with driver.session() as session:
            # 1. Get all node labels currently present in the DB
            labels_result = session.run("CALL db.labels()")
            labels = [row[0] for row in labels_result]
            
            dfs = {}
            node_pdf_map = {}
            
            # 2. Fetch data for each label
            for label in labels:
                logger.info(f"Fetching nodes for entity label: {label}")
                
                if label == 'Case':
                    query = f"MATCH (n:`{label}`) OPTIONAL MATCH (n)-[:HAS_DOCUMENT]->(d:Document) RETURN id(n) AS neo4j_node_id, n, collect(DISTINCT d.storage_id) AS pdf_paths"
                elif label == 'Document':
                    query = f"MATCH (n:Document) RETURN id(n) AS neo4j_node_id, n, [n.storage_id] AS pdf_paths"
                else:
                    query = f"MATCH (n:`{label}`) OPTIONAL MATCH (n)--(c:Case)-[:HAS_DOCUMENT]->(d:Document) RETURN id(n) AS neo4j_node_id, n, collect(DISTINCT d.storage_id) AS pdf_paths"
                
                results = session.run(query)
                
                records = []
                for record in results:
                    nid = record["neo4j_node_id"]
                    node = record["n"]
                    node_dict = dict(node)
                    node_dict = dict_to_string(node_dict)
                    
                    row_dict = {"neo4j_node_id": nid}
                    row_dict.update(node_dict)
                    
                    paths = [p for p in record["pdf_paths"] if p]
                    if paths:
                        rel_paths = []
                        for p in paths:
                            phys_path = resolve_physical_path(p)
                            if phys_path:
                                rel_paths.append(os.path.relpath(phys_path, PROJECT_ROOT))
                        if rel_paths:
                            rel_paths_str = "\n".join(rel_paths)
                            row_dict["pdf_paths"] = rel_paths_str
                            node_pdf_map[nid] = rel_paths_str
                        else:
                            row_dict["pdf_paths"] = ""
                            node_pdf_map[nid] = ""
                    else:
                        row_dict["pdf_paths"] = ""
                        node_pdf_map[nid] = ""
                        
                    records.append(row_dict)
                    
                if records:
                    df = pd.DataFrame(records)
                    # Sort to group identical pdf_paths for merging visually
                    if "pdf_paths" in df.columns:
                        df = df.sort_values(by=["pdf_paths", "neo4j_node_id"], na_position='last')
                    dfs[label] = df
                    
            # 3. Fetch all Relationships
            logger.info("Fetching relationships...")
            rel_query = """
            MATCH (a)-[r]->(b)
            RETURN id(a) as source_node_id,
                   coalesce(a.cnr, a.id, a.name_norm, "") as source_identifier,
                   labels(a)[0] as source_label, 
                   type(r) as relationship_type, 
                   id(b) as target_node_id,
                   coalesce(b.cnr, b.id, b.name_norm, "") as target_identifier,
                   labels(b)[0] as target_label,
                   properties(r) as rel_props
            """
            rel_results = session.run(rel_query)
            rel_records = []
            for record in rel_results:
                rec_dict = dict(record)
                props = rec_dict.pop('rel_props', {})
                if props:
                    props = dict_to_string(props)
                    for pk, pv in props.items():
                        rec_dict[f"rel_{pk}"] = pv
                        
                # Merge target and source PDF paths
                sid = rec_dict['source_node_id']
                tid = rec_dict['target_node_id']
                pset = set()
                
                if p1 := node_pdf_map.get(sid):
                    pset.update(p1.split('\n'))
                if p2 := node_pdf_map.get(tid):
                    pset.update(p2.split('\n'))
                
                rec_dict['pdf_paths'] = "\n".join(sorted([p for p in pset if p]))
                
                rel_records.append(rec_dict)
                
            if rel_records:
                df = pd.DataFrame(rel_records)
                if "pdf_paths" in df.columns:
                    df = df.sort_values(by=["pdf_paths", "source_node_id"], na_position='last')
                dfs['Relationships'] = df
                
    except Exception as e:
        logger.error(f"Error fetching from DB: {e}")
        return
    finally:
        driver.close()
        
    # 4. Save to Excel with merging logic
    if os.path.exists(OUTPUT_FILE):
        logger.info(f"Existing file {OUTPUT_FILE} found. Applying update-or-append logic...")
        try:
            existing_xls = pd.ExcelFile(OUTPUT_FILE)
            
            with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', mode='w') as writer:
                for sheet_name, new_df in dfs.items():
                    if sheet_name in existing_xls.sheet_names:
                        # Load existing sheet
                        existing_df = pd.read_excel(existing_xls, sheet_name=sheet_name)
                        
                        # Decide on a merge key: 'cnr', then 'neo4j_node_id', or nothing
                        merge_keys = [k for k in ['cnr', 'neo4j_node_id'] if k in new_df.columns and k in existing_df.columns]
                        
                        if merge_keys:
                            key = merge_keys[0]  # Prioritize 'cnr', then fallback to node_id
                            logger.info(f"Merging data in sheet '{sheet_name}' on key '{key}'...")
                            
                            # Set index for updating
                            existing_df.set_index(key, inplace=True)
                            new_df.set_index(key, inplace=True)
                            
                            # update existing
                            existing_df.update(new_df)
                            
                            # identify explicitly new rows
                            new_rows = new_df[~new_df.index.isin(existing_df.index)]
                            
                            final_df = pd.concat([existing_df, new_rows]).reset_index()
                        else:
                            # If no common matching key is found, just overwrite
                            logger.info(f"No valid merge key for sheet '{sheet_name}', overwriting...")
                            final_df = new_df
                            
                        # Dump to excel
                        final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    else:
                        # New sheet in neo4j that wasn't in excel yet
                        logger.info(f"Adding entirely new sheet '{sheet_name}'...")
                        new_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Copy over any sheets that were in Excel natively but weren't updated from DB
                for sheet_name in existing_xls.sheet_names:
                    if sheet_name not in dfs:
                        logger.info(f"Preserving un-updated sheet '{sheet_name}'...")
                        existing_df = pd.read_excel(existing_xls, sheet_name=sheet_name)
                        existing_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                # Make it look not like noodles
                format_excel_sheets(writer)
                
        except Exception as e:
            logger.error(f"Error during excel manipulation: {e}")
    else:
        logger.info(f"Creating new file {OUTPUT_FILE}...")
        try:
            with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
                for sheet_name, df in dfs.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Make it look not like noodles
                format_excel_sheets(writer)
        except Exception as e:
            logger.error(f"Error creating excel: {e}")
            
    logger.info("✅ Export and sync completed successfully!")

if __name__ == "__main__":
    export_all_to_excel()
